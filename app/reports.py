"""Report generation — the three output files (PRD 4.7, 4.8, 5.4).

Layout (matches the approved reference styling):
- Row 1: merged title banner — report name · Ref · Date · Patient · Generated
- Row 2: merged legend line (italic gray)
- Row 3: navy header row, white bold, wrapped
- Data rows: banded light green, thin borders, wrapped text, savings
  highlighting (🟢 >10%, 🟡 5–10%) layered on top, clickable URLs.

Price-match report: row selection logic UNCHANGED — exact four-criteria
matches only, cheaper than Schein only, sorted by total savings descending.
Match Type + Score render as one column: "EXACT (97%)".

Alternate purchase list: every reviewable scraped candidate that matches a
product NOT in the price-match report (plus equivalency-table findings).
Columns follow the approved reference sheet:
  Original Schein Product and Schein Price | Recommended Equivalent Product |
  Recommended Supplier | Price of the Equivalent | Product URL | Match Score |
  Equivalency Basis and Confidence Level | Estimated Savings vs. Schein Price
Tiers: EXACT = same product · CLOSE = compatible specs · POSSIBLE = needs review.
"""
from __future__ import annotations
import logging
import os
import re
import time
from pathlib import Path
from typing import List, Optional
from urllib.parse import quote_plus, urlparse


def _domain(url) -> str:
    try:
        d = urlparse(str(url or "")).netloc.lower()
        return d[4:] if d.startswith("www.") else d
    except Exception:
        return ""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import EquivalencyFinding, ItemResult, ParsedOrder, PriceCandidate

log = logging.getLogger(__name__)

# ------------------------------------------------------------- styling ------

NAVY_FILL = PatternFill("solid", fgColor="1F3650")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=12)
LEGEND_FONT = Font(italic=True, size=9, color="595959")
LINK_FONT = Font(color="0563C1", underline="single", size=9)
BAND_FILL = PatternFill("solid", fgColor="EAF3EA")     # light green banding
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")    # >10% savings
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")   # 5–10% savings
BOLD = Font(bold=True)
BOLD_RED = Font(bold=True, color="FF0000")
_thin = Side(style="thin", color="C9C9C9")
CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_green = Side(style="medium", color="538135")
TITLE_BORDER = Border(top=_green, bottom=_green)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
MONEY = '"$"#,##0.00'

PM_LEGEND = ("🟢 >10% savings   🟡 5–10% savings   EXACT = all four trust criteria "
             "confirmed (brand · product name · size/form · pack qty)   "
             "Product URL links to the verified product page")
ALT_LEGEND = ("Includes equivalency-table substitutions and all matching scraped "
              "prices not in the Price Match report  ·  EXACT = same product · "
              "CLOSE = compatible specs · POSSIBLE = needs review")


def _now() -> str:
    return time.strftime("%Y-%m-%d %H:%M")


def _dash(v) -> object:
    """'—' placeholder for empty cells (matches reference styling)."""
    if v is None:
        return "—"
    s = str(v).strip()
    return "—" if s == "" or s.lower() in ("none", "null", "n/a", "na", "-") else v


def _clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("none", "null", "n/a", "na", "-") else s


def _savings_pct(per_unit: float, unit_price: float):
    if not unit_price:
        return None
    return per_unit / unit_price * 100


def match_score(criteria: Optional[dict], confidence) -> int:
    """0–100: 60% weight on four-criteria coverage, 40% on AI confidence."""
    crit = criteria or {}
    met = sum(1 for k in ("brand_match", "name_match", "size_form_match", "pack_match")
              if crit.get(k))
    try:
        conf = int(confidence or 0)
    except (TypeError, ValueError):
        conf = 0
    return round(met / 4 * 60 + conf * 0.4)


def search_fallback_url(item) -> str:
    q = item.search_query or item.description
    return f"https://www.google.com/search?tbm=shop&q={quote_plus(q)}"


def _title_block(ws, title: str, legend: str, ncols: int,
                 headers: List[str], widths: List[int]):
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.font, c.alignment = TITLE_FONT, CENTER
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).border = TITLE_BORDER
    ws.row_dimensions[1].height = 26

    ws.append([legend])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1)
    c.font, c.alignment = LEGEND_FONT, CENTER
    ws.row_dimensions[2].height = 16

    ws.append(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=3, column=i)
        cell.fill, cell.font, cell.alignment = NAVY_FILL, HEADER_FONT, CENTER
        cell.border = CELL_BORDER
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"


def _style_row(ws, ridx: int, ncols: int, savings_pct: Optional[float],
               band: bool, wrap_cols: set[int], link_col: Optional[int] = None,
               url: str = "", bold_cols: set[int] = frozenset(),
               red_bold_cols: set[int] = frozenset()):
    fill = None
    if savings_pct is not None:
        fill = GREEN_FILL if savings_pct > 10 else (YELLOW_FILL if 5 <= savings_pct <= 10 else None)
    if fill is None and band:
        fill = BAND_FILL
    for col in range(1, ncols + 1):
        cell = ws.cell(row=ridx, column=col)
        cell.border = CELL_BORDER
        cell.alignment = WRAP if col in wrap_cols else Alignment(vertical="center")
        if fill is not None:
            cell.fill = fill
        if col in red_bold_cols:
            cell.font = BOLD_RED
        elif col in bold_cols:
            cell.font = BOLD
    if link_col and url:
        cell = ws.cell(row=ridx, column=link_col)
        cell.value = url
        cell.hyperlink = url
        cell.font = LINK_FONT
        cell.alignment = WRAP


def _money(ws, ridx: int, cols: List[int]):
    for c in cols:
        cell = ws.cell(row=ridx, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY


def _has_primary_row(r: ItemResult) -> bool:
    """An item is in price_match (excluded from alternate) iff it has at least
    one candidate passing the strict pack+variant+size gate."""
    if r.routed_to_alternate:
        return False
    return any(_pricematch_eligible(r.item, c) for c in r.candidates)


def _reviewable(c: PriceCandidate) -> bool:
    if c.price is None:
        return False
    reason = (c.rejected_reason or "").lower()
    return "login" not in reason and "sanity" not in reason


# tier mapping for the alternate sheet
_TIER = {
    "exact": ("EXACT", "EXACT — same product, different brand or channel; substitution is straightforward", 0),
    "approximate": ("APPROXIMATE", "CLOSE — same category, compatible specs, minor differences; substitution likely acceptable", 1),
    "rejected": ("POSSIBLE", "POSSIBLE — related product, requires client review before acting", 2),
    "unverified": ("POSSIBLE", "POSSIBLE — unverified listing, requires client review before acting", 3),
}
_CRIT_LABEL = {"brand_match": "brand", "name_match": "product name",
               "size_form_match": "form/spec", "pack_match": "pack size"}


def _basis_text(c: PriceCandidate) -> str:
    label, sentence, _ = _TIER.get(c.match_type, _TIER["unverified"])
    parts = [f"Web search (not in Price Match) · {label} ({match_score(c.criteria, c.confidence)}%)"]
    matched = [_CRIT_LABEL[k] for k, v in (c.criteria or {}).items() if v]
    missed = [_CRIT_LABEL[k] for k, v in (c.criteria or {}).items() if v is False]
    if matched:
        parts.append("Matching: " + ", ".join(matched))
    if missed:
        parts.append("Not matching: " + ", ".join(missed))
    note = _clean(c.notes) or _clean(c.rejected_reason)
    if note:
        parts.append(note)
    return sentence + "\n" + " · ".join(parts)


# ------------------------------------------------------ primary report ------

PM_HEADERS = ["Schein SKU", "Manufacturer Part\nNumber", "Description", "Qty\nOrder",
              "Schein Unit\nPrice", "Best Public Price\nFound", "Match Score",
              "Source Site", "Source Type", "Product URL Link",
              'Pack/Qty Condition\n(e.g. "6-pack price")',
              "Why Not Exact / Notes",
              "Schein\nPrice", "Best\nPrice",
              "Savings Per\nUnit", "Total Savings"]
PM_WIDTHS = [11, 15, 38, 7, 12, 13, 17, 17, 14, 42, 18, 40, 12, 12, 12, 12]
PM_WRAP = {3, 10, 11, 12}
OPTION_FONT = Font(italic=True, color="7F7F7F")

_OPT_RANK = {"exact": 0, "approximate": 1, "rejected": 2, "unverified": 3}


def _is_gated(c: PriceCandidate) -> bool:
    return "login" in (c.rejected_reason or "").lower()


def _pack_mismatch(item, c: PriceCandidate) -> bool:
    """Hard exclusion from the price-match report: a known different pack
    size is never negotiation-grade, no matter how cheap (PRD 4.5)."""
    if (c.criteria or {}).get("pack_match") is False:
        return True
    try:
        if c.pack_qty and item.pack_qty and int(c.pack_qty) != int(item.pack_qty):
            return True
    except (TypeError, ValueError):
        pass
    cond = (_clean(c.pack_condition) or "").lower()
    return any(k in cond for k in
               ("smaller pack", "larger pack", "case lot", "single unit",
                "instead of", "-pack price (ordered pack"))


def _brand_ok(item, c: PriceCandidate) -> bool:
    return bool((c.criteria or {}).get("brand_match")) or not getattr(item, "brand", None)


def _score_label(item, c: PriceCandidate) -> str:
    """Label derives from the criteria themselves, never from a raw
    match_type that contradicts them. Surfaces special tiers:
    GATED (login pricing), GENERIC EQUIVALENT (Q1: house-brand item, no exact
    competitor exists), VARIANT UNVERIFIED (Q2: variant not confirmed on page)."""
    score = match_score(c.criteria, c.confidence)
    if getattr(c, "out_of_stock", False):
        return f"OUT OF STOCK ({score}%)"
    if _is_gated(c):
        return f"GATED ({score}%)"
    # ISSUE 3: a price flagged unreliable (e.g. Groq read a number not on the page,
    # or it conflicts with the listing) must NEVER be presented as EXACT — that is
    # exactly the misleading "$26.93 EXACT" headline on a hallucinated price. Label
    # it UNVERIFIED PRICE so the buyer treats the number as needing confirmation.
    if getattr(c, "price_unreliable", False):
        return f"UNVERIFIED PRICE ({score}%)"
    if getattr(c, "is_generic_equivalent", False):
        return f"GENERIC EQUIVALENT ({score}%)"
    if _bulk_benefit(item, c):
        return f"BULK VALUE ({score}%)"
    # a login-rescued price with NO confirmed criteria (all None) reads oddly as
    # "POSSIBLE (0%)" — it's an unverified public price, so label it plainly
    if (c.match_type == "unverified" and not (c.criteria or {})
            and c.price is not None):
        return "UNVERIFIED"
    if _is_exact_cand(item, c):
        if getattr(c, "variant_unverified", False):
            return f"EXACT · VARIANT UNVERIFIED ({score}%)"
        return f"EXACT ({score}%)"
    if c.match_type in ("exact", "approximate"):
        return f"APPROXIMATE ({score}%)"
    return f"POSSIBLE ({score}%)"


def _is_exact_cand(item, c: PriceCandidate) -> bool:
    crit = c.criteria or {}
    return (crit.get("name_match") and crit.get("size_form_match")
            and crit.get("pack_match") and _brand_ok(item, c))


def _mismatch_reason(item, c: PriceCandidate) -> str:
    """Human-readable reason a candidate is in alternate, not price_match."""
    reasons = []
    crit = c.criteria or {}
    if crit.get("pack_match") is False or (
            c.pack_qty and item.pack_qty and int(c.pack_qty) != int(item.pack_qty)):
        pq = c.pack_qty if c.pack_qty else "?"
        reasons.append(f"PACK MISMATCH (page {pq} vs ordered {item.pack_qty or '?'})")
    if crit.get("size_form_match") is False:
        reasons.append("SIZE/VOLUME MISMATCH (e.g. 2L vs 5L) — verify on page")
    if getattr(c, "variant_unverified", False) or crit.get("name_match") is False:
        reasons.append(f"VARIANT MISMATCH (ordered: {item.variant or 'specified'})")
    if _is_gated(c):
        reasons.append("LOGIN-GATED price")
    if "page not found" in (c.rejected_reason or "").lower():
        reasons.append("DEAD/EXPIRED listing")
    return " · ".join(reasons)


def _bulk_benefit(item, c: PriceCandidate) -> bool:
    """A LARGER-pack listing that costs LESS in absolute terms than the ordered
    package is a strict win: more product for a lower total price (net32's Fuji
    50-pk EXPORT at $203.85 vs the ordered 48-pk at $372.66). Client rule
    2026-07-15: include these as price-match options despite the pack "mismatch".
    Same product only (MPN or name+size), in-stock, non-gated, reliable price.
    BRAND must also be confirmed or absent — a different-brand generic gutta
    percha (Meta Biomed vs SybronEndo) is not the same product."""
    if not (item.pack_qty and c.pack_qty and c.price and item.unit_price):
        return False
    if int(c.pack_qty) <= int(item.pack_qty):        # must be MORE than ordered
        return False
    if c.price >= item.unit_price:                    # and cheaper in absolute total
        return False
    if (_is_gated(c) or getattr(c, "out_of_stock", False)
            or getattr(c, "price_unreliable", False)
            or getattr(c, "variant_conflict", False)):
        return False
    if c.match_type == "unverified":
        return False
    crit = c.criteria or {}
    if not (getattr(c, "mpn_confirmed", False)
            or (crit.get("name_match") and crit.get("size_form_match"))):
        return False
    if (getattr(item, "brand", None) and not crit.get("brand_match")
            and not getattr(c, "mpn_confirmed", False)):
        return False
    return True


def _pricematch_eligible(item, c: PriceCandidate) -> bool:
    """STRICT price_match gate (client rule): pack quantity AND size/form AND
    variant must all match. Brand MAY differ (generic equivalent allowed, but
    labeled). Anything failing this goes to the alternate sheet instead.
    EXCEPTION: a larger pack at a lower total price (_bulk_benefit) is allowed."""
    crit = c.criteria or {}
    if c.price is None or c.scraped_product_name is None:
        return False
    # an unverified / page-unconfirmed price is never negotiation-grade EXACT:
    # it still appears as an option (with its ⚠ flag), but must not be presented
    # to the Schein rep as a confirmed four-criteria match.
    if getattr(c, "price_unreliable", False):
        return False
    # out-of-stock / no-longer-available listings are not a buyable price → never
    # negotiation-grade; they appear only as a clearly-labeled reference option.
    if getattr(c, "out_of_stock", False):
        return False
    if not (0.05 * item.unit_price <= c.price <= 3.0 * item.unit_price):
        return False
    if _is_gated(c):
        return False
    # BULK VALUE: a larger pack at a lower total price is a strict win — allow it
    # despite the pack "mismatch" (it still passes every other gate above).
    if _bulk_benefit(item, c):
        return True
    if _pack_mismatch(item, c):
        return False
    if "page not found" in (c.rejected_reason or "").lower():
        return False
    # a DETECTED wrong variant (wrong color/shade/size/flavor) is never price-match
    # grade — blocks even an MPN match (which can't co-exist with a real conflict).
    if getattr(c, "variant_conflict", False):
        return False
    # MPN match = identical product → price-match grade, overriding the free model's
    # unreliable name/size/pack criteria on aggregator-locked pages.
    if getattr(c, "mpn_confirmed", False):
        return True
    # variant must be confirmed (Q2: unverified variants are NOT price_match grade
    # under the strict rule — they drop to alternate)
    if getattr(c, "variant_unverified", False):
        return False
    # the three hard requirements: name/size + pack (+ size_form) all true
    if not (crit.get("name_match") and crit.get("size_form_match") and crit.get("pack_match")):
        return False
    return True


def _option_priority(item, c: PriceCandidate):
    """Sort key for the up-to-3 options shown per item. Lower sorts first.

    Client rule: among the options, give priority to candidates whose pack size
    AND variant match the ordered item — don't let a cheaper wrong-pack or
    wrong-variant listing outrank a correct one. Tiers (then cheapest within):
      0 — full price-match-eligible exact (pack + variant + size all confirmed)
      1 — pack matches AND variant confirmed
      2 — pack matches (variant unconfirmed/different)
      3 — everything else (pack mismatch / unknown)
    Price-unreliable candidates are pushed below their tier so a confirmed price
    is always preferred over an unverified one at the same match quality.
    """
    crit = c.criteria or {}
    pack_ok = not _pack_mismatch(item, c)
    variant_ok = (not getattr(c, "variant_unverified", False)
                  and crit.get("name_match") is not False)
    if _pricematch_eligible(item, c):
        tier = 0
    elif pack_ok and variant_ok:
        tier = 1
    elif pack_ok:
        tier = 2
    else:
        tier = 3
    unreliable = 1 if getattr(c, "price_unreliable", False) else 0
    oos = 1 if getattr(c, "out_of_stock", False) else 0
    # ISSUE 3 / out-of-stock: neither an unreliable price nor an out-of-stock
    # listing can sit in the top tiers (0/1), and an IN-STOCK option always
    # outranks an out-of-stock one — so a buyable price always headlines over an
    # unavailable one, even if the unavailable one is cheaper.
    if unreliable or oos:
        tier = max(tier, 2)
    return (tier, oos, unreliable, c.price if c.price is not None else 1e9)


def _poolable(item, c: PriceCandidate) -> bool:
    """A candidate eligible to be shown as a price-match OPTION: priced within the
    sane band, page-verified, not login-gated, not a dead listing."""
    unit = item.unit_price
    # SAVINGS-ONLY price_match (client QA): a candidate priced AT or ABOVE the
    # Schein price is not a saving — it belongs in the alternate sheet as a
    # reference, never in the price_match report. Disable via PRICEMATCH_SAVINGS_ONLY=0.
    savings_only = os.environ.get("PRICEMATCH_SAVINGS_ONLY", "1") not in ("0", "false", "False")
    # A confirmed EXACT/APPROXIMATE (or MPN-confirmed) match keeps the wide sanity
    # band — a genuine deep discount is allowed. But a WEAK match (rejected/possible/
    # unverified) that is implausibly cheap is almost always the WRONG product or a
    # smaller pack (Oral-B: $6.20 single Healthy Gums / $11.25 trial vs a $50.99
    # 72-count box) — require it to be at least 30% of the Schein unit price before
    # it can pose as a price-match option, else it goes to the alternate sheet.
    strong = c.match_type in ("exact", "approximate") or getattr(c, "mpn_confirmed", False)
    floor = (0.05 if strong else 0.30) * unit
    return (c.price is not None
            and floor <= c.price <= 3.0 * unit
            and (not savings_only or c.price < unit)       # must BEAT Schein — no-saving rows → alternate
            and c.scraped_product_name is not None
            and not _is_gated(c)
            and not getattr(c, "out_of_stock", False)      # OOS → alternate sheet, never a price-match option
            and not getattr(c, "variant_conflict", False)  # wrong color/shade/size/flavor → alternate, never price-match
            and not getattr(c, "pack_conflict", False)     # different pack size → not a like-for-like price → alternate
            # a REJECTED candidate is one the pipeline decided is NOT the ordered
            # product (net32's Defend 830L-012 bur for a Meisinger 841G-012 order,
            # "Product name and MPN do not match") — it must never be a price-match
            # option. Login/OOS rejections are rescued to 'unverified' earlier, so
            # what remains as 'rejected' is a genuine product mismatch.
            and c.match_type != "rejected"
            # an UNVERIFIED candidate has NO confirmed match criteria — it's a
            # login-rescued structured price whose product identity was never
            # validated.  Showing it as a price-match option risks headlining a
            # wrong product (Safco G2-Bond at $68.99 for an OptiBond order).
            # Route to alternate sheet; verified matches headline instead.
            and c.match_type != "unverified"
            and "page not found" not in (c.rejected_reason or "").lower())


def _select_options(r: ItemResult, options_per_item: int = 3) -> list:
    """The up-to-N candidates shown for this item in the Price Match report:
    cheapest price-match-eligible EXACT first, then the next best by pack/variant
    priority. Shared by the price-match writer (to render) and the alternate
    writer (to EXCLUDE these so they aren't repeated)."""
    # Dedup by URL AND by supplier domain: a single supplier should appear once
    # (its best/cheapest listing), so the same store can't fill multiple option
    # slots (stardentalsupplies showed twice for the Teal tips). Candidates are
    # sorted by priority first, so the kept one per domain is the best.
    pool, seen_u, seen_dom = [], set(), set()
    for c in sorted(r.candidates, key=lambda c: _option_priority(r.item, c)):
        d = _domain(c.url)
        if not _poolable(r.item, c) or c.url in seen_u or d in seen_dom:
            continue
        seen_u.add(c.url); seen_dom.add(d)
        pool.append(c)
    if not pool:
        return []
    exacts = [c for c in pool if _pricematch_eligible(r.item, c)]
    opts = []
    if exacts:
        opts.append(min(exacts, key=lambda c: c.price))   # cheapest eligible exact
    for c in pool:
        if len(opts) >= options_per_item:
            break
        if any(c is o for o in opts):
            continue
        opts.append(c)
    return opts


# -------------------------------- equivalency rows (cross-brand alternatives) --

EQUIV_FILL = PatternFill("solid", fgColor="F3E8FF")     # light purple
EQUIV_FONT = Font(bold=True, color="6B21A8")            # dark purple label


def _find_equivalents(item, candidates, shown_urls: set,
                      max_results: int = 2) -> list:
    """Find cross-brand equivalent candidates cheaper than Schein.
    ONLY used when no main options exist (NO SUPPLIER MATCH / REFERENCE).
    Does NOT change any existing matching logic — purely additive surfacing."""
    ordered_pack = getattr(item, "pack_qty", None)
    equivs = []
    seen_domains = set()
    for c in candidates:
        if c.url in shown_urls:
            continue
        if c.price is None or c.price >= item.unit_price:
            continue
        if getattr(c, "out_of_stock", False):
            continue
        if getattr(c, "variant_conflict", False):
            continue
        if getattr(c, "pack_conflict", False):
            continue
        if getattr(c, "price_unreliable", False):
            continue
        if c.scraped_product_name is None:
            continue
        if not (0.05 * item.unit_price <= c.price <= item.unit_price):
            continue
        if _is_gated(c):
            continue
        # pack must match if both are known
        if ordered_pack and c.pack_qty and c.pack_qty != ordered_pack:
            continue
        dom = _domain(c.url)
        if dom in seen_domains:
            continue
        seen_domains.add(dom)
        equivs.append(c)
    equivs.sort(key=lambda c: c.price)
    return equivs[:max_results]


def _write_equivalent_rows(ws, r: ItemResult) -> None:
    """Render cross-brand equivalent rows for items with no main options.
    Clearly labeled so the user knows to verify substitutability."""
    item = r.item
    shown = {c.url for c in (r.candidates if hasattr(r, "_shown_opts") else [])
             if getattr(c, "_shown", False)}
    equivs = _find_equivalents(item, r.candidates, shown)
    if not equivs:
        return
    for c in equivs:
        per_unit = round(item.unit_price - c.price, 2)
        total = round(per_unit * item.qty, 2)
        pack_note = ""
        if c.pack_qty is None:
            pack_note = " · pack size not confirmed on page — verify before ordering"
        name_short = (c.scraped_product_name or c.title or "")[:60]
        note = (f"EQUIVALENT PRODUCT — different brand, same product type "
                f"({name_short}). Verify clinical equivalence before "
                f"substituting{pack_note}.")
        ws.append([item.schein_sku, _dash(item.mpn),
                   f"   ↳ ⚡ Equivalent — {item.description}",
                   item.qty, item.unit_price, c.price,
                   f"EQUIVALENT ({c.confidence}%)" if c.confidence else "EQUIVALENT",
                   c.source_site, _source_type_for(c), c.url,
                   _dash(_clean(c.pack_condition)), note,
                   item.unit_price, c.price, per_unit, total])
        ridx = ws.max_row
        _style_row(ws, ridx, len(PM_HEADERS), None, False, PM_WRAP,
                   link_col=10, url=c.url, red_bold_cols={14})
        _money(ws, ridx, [5, 6, 13, 14, 15, 16])
        for col in range(1, len(PM_HEADERS) + 1):
            ws.cell(row=ridx, column=col).fill = EQUIV_FILL
        ws.cell(row=ridx, column=3).font = EQUIV_FONT
        if c.url:
            ws.cell(row=ridx, column=10).font = LINK_FONT
        ws.row_dimensions[ridx].height = 50


# ------------------------------------------- marketplace rows (🅐/🅦) --------
# Dedicated rows per item — Amazon, Walmart — appended after the regular
# options. A marketplace PRICE is shown only under the strict client rule:
# same product AND same pack/size verified on the listing page (brand must
# match too, or the MPN must appear on the listing). Anything less renders
# as a "not on <marketplace>" row with the reason.

MARKETPLACES = [("amazon", "🅐 Amazon", "AMAZON"),
                ("walmart", "🅦 Walmart", "WALMART")]

MKT_STYLE = {
    "amazon":  (PatternFill("solid", fgColor="FFE8CC"), Font(bold=True, color="C45500")),
    "walmart": (PatternFill("solid", fgColor="DCEBFB"), Font(bold=True, color="0071CE")),
}


def _marketplace_eligible(item, c: PriceCandidate) -> bool:
    return (_pricematch_eligible(item, c)
            and (_brand_ok(item, c) or getattr(c, "mpn_confirmed", False)))


def _is_house_brand_item(item) -> bool:
    """A Henry Schein house-brand order (Premium/Criterion/Acclean/Maxima…) has
    NO exact-brand competitor anywhere — every marketplace listing is a generic
    by definition, so the strict brand gate would render 'not found' forever."""
    return "henry schein" in (getattr(item, "brand", "") or "").lower()


_TYPE_STOP = {"the", "and", "for", "with", "non", "sterile", "premium", "brand",
              "dental", "dentistry", "absorbable", "disposable", "henry", "schein"}


def _generic_marketplace_eligible(item, c: PriceCandidate) -> bool:
    """Relaxed gate for the HOUSE-BRAND generic marketplace row ONLY: a generic
    has a DIFFERENT product name by definition, so name/brand match is not
    required. Instead demand (a) the ordered pack confirmed on the listing,
    (b) strong product-TYPE token overlap (cotton+rolls+#2+2000), (c) a sane
    in-stock, non-gated, reliable price. Keeps a random cheap product from posing
    as the generic while letting the true same-pack alternative through."""
    # a generic listing the LLM rejected for brand often has scraped_product_name
    # nulled — fall back to the listing title / structured name for identity
    name_src = (c.scraped_product_name or c.title or getattr(c, "structured_name", None) or "")
    if c.price is None or not name_src:
        return False
    if _is_gated(c) or getattr(c, "out_of_stock", False) or getattr(c, "price_unreliable", False):
        return False
    if not (0.05 * item.unit_price <= c.price <= 3.0 * item.unit_price):
        return False
    cpack = c.pack_qty
    if cpack is None and item.pack_qty:      # derive from title/name/url ("2000/Bx", "2000-Bx")
        hay = f"{name_src} {c.url or ''}"
        m = re.search(r"\b(\d{2,5})\s*[-/]?\s*(?:bx|box|pk|pack|ct|count|ca|case)\b", hay, re.I)
        if m:
            cpack = int(m.group(1))
    if not (item.pack_qty and cpack and int(cpack) == int(item.pack_qty)):
        return False
    # NOTE: don't trust criteria['size_form_match'] here — when the LLM rejects a
    # generic on brand it nulls ALL criteria to False, which would wrongly block a
    # correct-size generic. Size is instead discriminated by the exact pack match
    # above and the ordered spec tokens ("#2") in the overlap check below.
    page = f"{name_src} {c.title or ''}".lower()
    # a size/grade marker in the order ("#2", "#4") is REQUIRED verbatim — it is
    # the one token that separates a #2 cotton roll from a #4, and token-overlap
    # alone dilutes it. A conflicting/absent grade disqualifies the generic.
    grades = re.findall(r"#\s*\d+", (item.description or "").lower())
    if grades and not all(re.sub(r"\s", "", g) in re.sub(r"\s", "", page) for g in grades):
        return False
    ref = re.findall(r"[a-z0-9#]+", (item.description or "").lower())
    toks = [w for w in ref if len(w) > 1 and w not in _TYPE_STOP]
    # the ALPHA product-noun tokens (cotton, rolls) carry the identity — a gauze
    # sponge shares the generic numeric tokens (#2, 2000) but none of these, so
    # require a strong match on the words specifically, not just overall overlap.
    words = [w for w in toks if w.isalpha()]
    if words and sum(1 for w in words if w in page) / len(words) < 0.6:
        return False
    if not toks:
        return False
    return sum(1 for w in toks if w in page) / len(toks) >= 0.55


def _pick_marketplace(item, mcands: List[PriceCandidate], key: str):
    """(best_candidate, reason) for one marketplace: the cheapest listing passing
    the strict same-product+pack gate, else (None, why-not). Client rule
    2026-07-15: for HOUSE-BRAND items only, a same-pack generic may show,
    explicitly labelled — reason sentinel "generic"."""
    pool = [c for c in mcands if getattr(c, "marketplace", None) == key]
    eligible = [c for c in pool if _marketplace_eligible(item, c)]
    if eligible:
        # a SEEDED (operator-vouched) eligible listing wins over an ambiguous
        # search hit at the same/higher price — the seeded eBay A3 (Ref 000140)
        # must beat a shadeless "Fuji II Gold Label" $180 that merely ties on price.
        seeded = [c for c in eligible if getattr(c, "_seed", False)]
        if seeded:
            return min(seeded, key=lambda c: c.price), ""
        return min(eligible, key=lambda c: c.price), ""
    if _is_house_brand_item(item):
        generics = [c for c in pool if _pricematch_eligible(item, c)
                    or _generic_marketplace_eligible(item, c)]
        if generics:
            return min(generics, key=lambda c: c.price), "generic"
    if not pool:
        return None, "no listing surfaced in a site-restricted search"
    # explain the closest miss so the buyer knows what WAS there
    scraped = [c for c in pool if c.scraped_product_name or c.rejected_reason]
    if not scraped:
        return None, "listings found but none could be page-verified this run"
    best = max(scraped, key=lambda c: match_score(c.criteria, c.confidence))
    why = (_mismatch_reason(item, best) or _clean(best.rejected_reason)
           or "closest listing failed the same-product + same-pack check")
    return None, f"closest listing did not qualify: {why}"


def _source_type_for(c, *, marketplace: str | None = None) -> str:
    from .admin_config import classify_source_type
    return classify_source_type(
        getattr(c, "url", None) or getattr(c, "source_site", None),
        marketplace=marketplace or getattr(c, "marketplace", None),
        is_generic=bool(getattr(c, "is_generic_equivalent", False)),
    )


def _write_marketplace_rows(ws, r: ItemResult) -> None:
    """The 🅐/🅦 rows for one item group."""
    item = r.item
    for key, label, score_label in MARKETPLACES:
        name = label.split(" ", 1)[1]
        best, why = _pick_marketplace(item, r.marketplace_candidates or [], key)
        src_type = "Marketplace"
        if best is not None:
            per_unit = round(item.unit_price - best.price, 2)
            total = round(per_unit * item.qty, 2)
            generic = why == "generic"
            if generic:
                src_type = "Generic/Equivalent"
            if best.price < item.unit_price:
                row_label, sp, st = f"   {label}", per_unit, total
                note = (f"{name} price — ${per_unit:,.2f}/unit below Schein · same "
                        f"product & pack verified on the listing page")
            else:
                row_label, sp, st = f"   {label} (reference)", "", ""
                note = (f"{name} reference price — ${abs(per_unit):,.2f}/unit ABOVE "
                        f"Schein (shown for reference) · same product & pack verified")
            if generic:
                row_label += " (generic)"
                note = (f"GENERIC EQUIVALENT — verify before substituting · house-brand "
                        f"order, no exact-brand listing can exist on {name}; this is the "
                        f"best same-pack generic ({(best.scraped_product_name or best.title or '')[:60]}) · "
                        + note)
            extra = _clean(best.notes)
            if extra:
                note += f" · {extra}"
            # Schein price / qty carried on every marketplace option row
            ws.append([item.schein_sku, _dash(item.mpn), row_label, item.qty, item.unit_price,
                       best.price, score_label, best.source_site, src_type, best.url,
                       _dash(_clean(best.pack_condition)), note,
                       item.unit_price, best.price, sp, st])
            ridx = ws.max_row
            _style_row(ws, ridx, len(PM_HEADERS), None, False, PM_WRAP,
                       link_col=10, url=best.url, red_bold_cols={14})
            _money(ws, ridx, [5, 6, 13, 14, 15, 16])
        else:
            ws.append([item.schein_sku, _dash(item.mpn), f"   {label}", item.qty, item.unit_price,
                       f"not on {name}", score_label, "—", src_type, "", "—",
                       f"No matching {name} product found — {why}",
                       item.unit_price, "", "", ""])
            ridx = ws.max_row
            _style_row(ws, ridx, len(PM_HEADERS), None, False, PM_WRAP)
            _money(ws, ridx, [5, 13])
        fill, brand_font = MKT_STYLE[key]
        for col in range(1, len(PM_HEADERS) + 1):
            ws.cell(row=ridx, column=col).fill = fill
        ws.cell(row=ridx, column=3).font = brand_font
        if best is not None and best.url:
            ws.cell(row=ridx, column=10).font = LINK_FONT
        ws.row_dimensions[ridx].height = 44


def _reference_option(item, candidates):
    """Cheapest page-verified EXACT match (all four criteria at the ordered pack)
    that is priced AT OR ABOVE Schein — i.e. a genuine same-product listing the
    savings-only rule excludes from the options. Shown as a labelled "(reference)"
    main row instead of a bare NO SUPPLIER MATCH, so an item WITH exact matches
    (just none cheaper) is never reported as if nothing was found. In-stock,
    non-gated, price-reliable only. Returns the candidate or None."""
    pool = [c for c in candidates
            if _pricematch_eligible(item, c)
            and c.price is not None and c.price >= item.unit_price
            and not getattr(c, "out_of_stock", False)
            and not getattr(c, "price_unreliable", False)
            and not _is_gated(c)]
    return min(pool, key=lambda c: c.price) if pool else None


def _oos_reference(item, candidates, shown_lo):
    """Cheapest OUT-OF-STOCK candidate that confirms all four criteria at the
    ordered pack AND undercuts the best in-stock option shown (client QA:
    carolinadental's $245.15 Fuji A3 vs the $318.96 headline). Not a buyable
    price, so it renders as a clearly-labelled reference sub-row — availability
    is the ONLY gate waived; pack/criteria/price-trust gates all still apply.
    Returns the candidate or None."""
    pool = [c for c in candidates
            if getattr(c, "out_of_stock", False) and c.price is not None
            and _is_exact_cand(item, c) and not _pack_mismatch(item, c)
            and not getattr(c, "price_unreliable", False)
            and not getattr(c, "variant_conflict", False)
            and 0.05 * item.unit_price <= c.price <= 3.0 * item.unit_price]
    if not pool:
        return None
    c = min(pool, key=lambda x: x.price)
    return c if (shown_lo is None or c.price < shown_lo) else None


def write_price_match_report(order: ParsedOrder, results: List[ItemResult],
                             out: Path, options_per_item: int = 3) -> Path:
    """Primary negotiation report — up to 3 options per item.
    Row layout (per approved screenshot): the item's main row carries the full
    item details with Option 1; Options 2-3 render as indented "↳ Option N"
    sub-rows (SKU/MPN/Qty/Schein-price left blank) directly beneath it.
    Item groups are sorted by their best total savings descending."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Match"
    title = (f"Henry Schein Price Match Negotiation  ·  Ref: {order.reference or '—'}"
             f"  ·  Date: {order.order_date or '—'}"
             f"  ·  Patient: {getattr(order, 'ship_to_name', None) or '—'}"
             f"  ·  Generated: {_now()}")
    legend = ("🟢 >10% savings   🟡 5–10% savings   Main row = best option (EXACT when "
              "available) · ↳ Option 2-3 = next-closest matches with reasoning · "
              "⚡ EQUIVALENT = different brand, same product type (verify before substituting) · "
              "🅐/🅦 = Amazon/Walmart check (price shown only when the same "
              "product AND pack is verified) · GATED = login pricing, verify manually")
    _title_block(ws, title, legend, len(PM_HEADERS), PM_HEADERS, PM_WIDTHS)

    groups = []
    for r in results:
        # Options shown for this item (cheapest eligible exact, then next best by
        # pack/variant priority). Same selection the alternate writer uses to
        # exclude these rows so they aren't repeated there. EVERY item gets a
        # group (client rule): items with no supplier match render a placeholder
        # main row so their Amazon/Walmart rows still show; they sort last.
        opts = _select_options(r, options_per_item)
        if opts:
            best_total = max(round((r.item.unit_price - c.price) * r.item.qty, 2)
                             for c in opts)
        else:
            best_total = float("-inf")
        groups.append((best_total, r, opts))

    groups.sort(key=lambda g: g[0], reverse=True)
    band = 0
    for _, r, opts in groups:
        if not opts:
            ref = _reference_option(r.item, r.candidates)
            if ref is not None:
                # exact match(es) exist but none beat Schein — show the cheapest
                # as a clearly-labelled reference row instead of NO SUPPLIER MATCH
                over = round(ref.price - r.item.unit_price, 2)
                ws.append([r.item.schein_sku, _dash(r.item.mpn), r.item.description,
                           r.item.qty, r.item.unit_price, ref.price,
                           f"REFERENCE — {_score_label(r.item, ref)}",
                           ref.source_site, _source_type_for(ref), ref.url,
                           _dash(_clean(ref.pack_condition)),
                           f"NO CHEAPER SUPPLIER — closest exact match is ${ref.price:,.2f} "
                           f"(${over:,.2f}/unit ABOVE Schein ${r.item.unit_price:,.2f}); shown "
                           f"for reference. Schein is already competitive here.",
                           r.item.unit_price, ref.price, "", ""])
                ridx = ws.max_row
                _style_row(ws, ridx, len(PM_HEADERS), None, band % 2 == 0, PM_WRAP,
                           link_col=10, url=ref.url, red_bold_cols={14})
                _money(ws, ridx, [5, 6, 13, 14])
                ws.row_dimensions[ridx].height = 48
            else:
                # placeholder main row — item had no verifiable public supplier match
                ws.append([r.item.schein_sku, _dash(r.item.mpn), r.item.description,
                           r.item.qty, r.item.unit_price, "—", "NO SUPPLIER MATCH",
                           "—", "Other", search_fallback_url(r.item), "—",
                           "No public supplier listing passed verification this run — "
                           "see the Alternate Purchases sheet for near-matches.",
                           r.item.unit_price, "—", "", ""])
                ridx = ws.max_row
                _style_row(ws, ridx, len(PM_HEADERS), None, band % 2 == 0, PM_WRAP,
                           link_col=10, url=search_fallback_url(r.item))
                _money(ws, ridx, [5, 13])
            ws.row_dimensions[ridx].height = 44
        for n, c in enumerate(opts, start=1):
            per_unit = round(r.item.unit_price - c.price, 2)
            total = round(per_unit * r.item.qty, 2)
            pct = _savings_pct(per_unit, r.item.unit_price)
            # explicit flags the client must see at a glance
            flags = []
            if c.price >= r.item.unit_price:
                flags.append(f"⚠ NO SAVING — ${c.price:,.2f} is HIGHER than Schein "
                             f"${r.item.unit_price:,.2f}; shown as closest match only")
            if getattr(c, "price_unreliable", False):
                flags.append("⚠ PRICE UNRELIABLE — auto-extracted price conflicts with "
                             "the listing; confirm on the page before using")
            if getattr(c, "out_of_stock", False):
                flags.append("⚠ OUT OF STOCK — this listing is no longer available; "
                             "not a buyable price, shown for reference only")
            flag_prefix = (" · ".join(flags) + " · ") if flags else ""
            if _bulk_benefit(r.item, c):
                iu = r.item.unit_price / r.item.pack_qty
                cu = c.price / c.pack_qty
                reason = (f"BULK VALUE — {c.pack_qty}/pk (MORE than the ordered "
                          f"{r.item.pack_qty}/pk) at a LOWER total price: ${cu:,.2f}/unit "
                          f"vs Schein ${iu:,.2f}/unit. More product for less money — "
                          f"verify it's the same item before ordering.")
            elif getattr(c, "is_generic_equivalent", False):
                reason = ("GENERIC EQUIVALENT — same product type, different/no brand; "
                          "no competitor sells the exact Schein house-brand item. "
                          "Verify clinical equivalence before substituting.")
            elif _is_exact_cand(r.item, c):
                base = "Exact — all four trust criteria confirmed"
                if getattr(c, "variant_unverified", False):
                    base += (" · VARIANT UNVERIFIED — page did not confirm the ordered "
                             f"variant ({r.item.variant or 'specified variant'}); verify before buying")
                if n > 1:
                    base += " (alternate source, higher price than Option 1)"
                reason = base
            else:
                crit = c.criteria or {}
                matched = [_CRIT_LABEL[k] for k, v in crit.items() if v]
                not_matched = [_CRIT_LABEL[k] for k, v in crit.items() if v is False]
                mm = _mismatch_reason(r.item, c)
                parts = []
                if mm:
                    parts.append(mm)
                if matched:
                    parts.append("Matching: " + ", ".join(matched))
                if not_matched:
                    parts.append("Not matching: " + ", ".join(not_matched))
                note = _clean(c.notes) or _clean(c.rejected_reason)
                if note:
                    parts.append(note)
                reason = " · ".join(parts) or "Not exact — could not confirm all four criteria"

            reason = flag_prefix + reason   # prepend ⚠ NO SAVING / PRICE UNRELIABLE flags
            src_type = _source_type_for(c)

            # Client request: Schein unit price (and identifying cols) carry onto
            # every option / match row for side-by-side comparison.
            if n == 1:
                desc = r.item.description
            else:
                desc = f"   ↳ Option {n} — {r.item.description}"
            ws.append([r.item.schein_sku, _dash(r.item.mpn), desc,
                       r.item.qty, r.item.unit_price, c.price, _score_label(r.item, c),
                       c.source_site, src_type, c.url,
                       _dash(_clean(c.pack_condition)), reason,
                       r.item.unit_price, c.price, per_unit, total])
            ridx = ws.max_row
            if n == 1:
                _style_row(ws, ridx, len(PM_HEADERS), pct, band % 2 == 0, PM_WRAP,
                           link_col=10, url=c.url, bold_cols={16}, red_bold_cols={14})
            else:
                _style_row(ws, ridx, len(PM_HEADERS), None, False, PM_WRAP,
                           link_col=10, url=c.url, red_bold_cols={14})
                ws.cell(row=ridx, column=3).font = OPTION_FONT
            _money(ws, ridx, [5, 6, 13, 14, 15, 16])
            ws.row_dimensions[ridx].height = 56

        # Labelled backorder/long-lead row: net32 flips a seller between
        # "Backordered" and "Long Handling Time" depending on the scrape's
        # delivery location, so the genuine lowest (e.g. XLight Shine $26.73)
        # would otherwise appear or vanish run-to-run. Surface the cheapest
        # excluded seller — when it undercuts the best shown option — as a clearly
        # labelled row so the buyer always sees it WITH its caveat and decides.
        bo = []
        for c in r.candidates:
            for o in (getattr(c, "backorder_options", None) or []):
                if o.get("price"):
                    bo.append((o, c))
        if bo and opts:
            shown_lo = min(c.price for c in opts)
            o, c = min(bo, key=lambda t: t[0]["price"])
            if o["price"] < shown_lo:
                vend = o.get("vendor") or "seller"
                status = o.get("status") or "Backordered"
                per_unit = round(r.item.unit_price - o["price"], 2)
                total = round(per_unit * r.item.qty, 2)
                reason = (f"⚠ {status.upper()} — ${o['price']:,.2f} via {vend} on "
                          f"{c.source_site} is the lowest listing but is NOT in stock "
                          f"(net32 stock varies by location; verify before relying on it). "
                          f"Shown for reference below the in-stock options.")
                ws.append([r.item.schein_sku, _dash(r.item.mpn),
                           f"   ↳ ⚠ Backorder — {r.item.description}",
                           r.item.qty, r.item.unit_price, o["price"],
                           "BACKORDER", c.source_site, _source_type_for(c), c.url,
                           "—", reason,
                           r.item.unit_price, o["price"], per_unit, total])
                ridx = ws.max_row
                _style_row(ws, ridx, len(PM_HEADERS), None, False, PM_WRAP,
                           link_col=10, url=c.url, red_bold_cols={14})
                ws.cell(row=ridx, column=3).font = OPTION_FONT
                _money(ws, ridx, [5, 6, 13, 14, 15, 16])
                ws.row_dimensions[ridx].height = 56

        # (Out-of-stock listings are excluded from price_match entirely per client
        # rule 2026-07-15 — no OOS reference row. OOS candidates already never pool
        # or headline; they remain in the Alternate Purchases / Evidence sheets.)

        if not opts:
            _write_equivalent_rows(ws, r)

        # 🅐/🅦 marketplace rows — always rendered, found or not
        _write_marketplace_rows(ws, r)
        band += 1
    wb.save(out)
    n_rows = sum(len(o) for _, _, o in groups)
    log.info("Price match report: %d item(s), %d option row(s) + %d marketplace "
             "row(s) → %s", len(groups), n_rows, len(MARKETPLACES) * len(groups), out.name)
    return out


# ---------------------------------------------------- alternate report ------

ALT_HEADERS = ["Original Schein Product", "Schein Unit\nPrice",
               "Recommended Equivalent Product", "Recommended Supplier",
               "Source Type", "Price of the\nEquivalent", "Product URL", "Match Score",
               "Equivalency Basis and Confidence Level",
               "Estimated Savings vs.\nSchein Price"]
ALT_WIDTHS = [34, 12, 36, 19, 14, 12, 44, 16, 48, 18]
ALT_WRAP = {1, 3, 9}


def _alt_row(ws, band, item, equiv_name, supplier, price, url, score_label,
             basis, savings_cell, pct, source_type: str = "Other"):
    ws.append([
        f"{item.description}\nSKU: {item.schein_sku}",
        item.unit_price,
        equiv_name, supplier or "—", source_type, price, url, score_label, basis, savings_cell,
    ])
    ridx = ws.max_row
    _style_row(ws, ridx, len(ALT_HEADERS), pct, band % 2 == 0, ALT_WRAP,
               link_col=7, url=url)
    _money(ws, ridx, [2, 6, 10])
    ws.row_dimensions[ridx].height = 64
    return ridx


def write_alternate_purchase_list(order: ParsedOrder,
                                  findings: List[EquivalencyFinding], out: Path,
                                  results: Optional[List[ItemResult]] = None) -> Path:
    """Stage 2 output — equivalency-table findings PLUS every reviewable
    scraped candidate for items that did not make the Price Match report.
    Items here never appear in the Price Match report (stream separation)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alternate Purchases"
    title = (f"Alternate Purchase Recommendations  ·  Ref: {order.reference or '—'}"
             f"  ·  Generated: {_now()}"
             f"  ·  Items here do NOT appear in the Price Match report")
    _title_block(ws, title, ALT_LEGEND, len(ALT_HEADERS), ALT_HEADERS, ALT_WIDTHS)

    entries = []   # (tier_rank, item_order_idx, price, row-args)
    order_idx = {i.schein_sku: n for n, i in enumerate(order.items)}

    # Candidates already SHOWN as options in the Price Match report — exclude them
    # here so the alternate sheet never repeats a row that's in price_match.
    shown_in_pm = {(r.item.schein_sku, c.url)
                   for r in (results or []) for c in _select_options(r)}
    for r in (results or []):
        opts = _select_options(r)
        if not opts:
            rc = _reference_option(r.item, r.candidates)
            if rc is not None:
                shown_in_pm.add((r.item.schein_sku, rc.url))

    # A — equivalency-table findings (confirmed substitutions)
    equivalency_skus = set()
    for f in findings:
        if f.confidence_level not in ("exact_equivalent", "close_equivalent"):
            continue
        equivalency_skus.add(f.item.schein_sku)
        label = "EXACT" if f.confidence_level == "exact_equivalent" else "CLOSE"
        sentence = _TIER["exact"][1] if label == "EXACT" else _TIER["approximate"][1]
        basis = sentence + "\nEquivalency table (client-maintained) · " + (f.basis or "")
        per_unit = round(f.item.unit_price - f.price, 2) if f.price else None
        pct = _savings_pct(per_unit, f.item.unit_price) if per_unit is not None else None
        savings = (per_unit if per_unit and per_unit > 0
                   else (f"Equiv ${f.price:,.2f} > Schein ${f.item.unit_price:,.2f}"
                         if f.price else "—"))
        entries.append((-1, order_idx.get(f.item.schein_sku, 999), f.price or 0,
                        (f.item, f.equivalent_name, f.supplier, f.price,
                         f.url or search_fallback_url(f.item), f"{label} (table)",
                         basis, savings, pct, "Generic/Equivalent")))

    # B — EVERY reviewable candidate that is NOT price_match-eligible, for ALL
    #     items. The alternate sheet now covers every product: pack/variant/size
    #     mismatches (even from items that DO have a price_match row), plus all
    #     candidates from items with no price_match row at all. Nothing dropped.
    for r in (results or []):
        if r.item.schein_sku in equivalency_skus:
            continue
        pool, discovered, seen = [], [], set()
        for c in r.candidates:
            if c.url in seen:
                continue
            seen.add(c.url)
            if (r.item.schein_sku, c.url) in shown_in_pm:
                continue                  # already shown in the Price Match report
            if _reviewable(c):
                pool.append(c)
            elif c.price is None and "login" not in (c.rejected_reason or "").lower():
                # Discovered during search but never priced this run — almost
                # always the per-item scrape cap (only the first N of the
                # discovered URLs are fetched). Surface them so the alternate
                # sheet is a TRUE catch-all of every link not in the Price Match
                # report. Login-gated URLs are deliberately left to the Flagged
                # Sites sheet instead of cluttering this one.
                discovered.append(c)
        if not pool and not discovered:
            # genuinely nothing found — placeholder so the item is never invisible
            if not _has_primary_row(r):
                url = search_fallback_url(r.item)
                entries.append((9, order_idx.get(r.item.schein_sku, 999), 0,
                                (r.item, "— no public candidate found this run —", "—",
                                 None, url, "POSSIBLE (0%)",
                                 "POSSIBLE — no usable public listing found; link opens a supplier search",
                                 "—", None, "Other")))
            continue
        for c in pool:
            mismatch = _mismatch_reason(r.item, c)
            label, _, rank = _TIER.get(c.match_type, _TIER["unverified"])
            if c.price is None:
                savings, pct = "—", None
            else:
                per_unit = round(r.item.unit_price - c.price, 2)
                pct = _savings_pct(per_unit, r.item.unit_price)
                savings = (per_unit if per_unit > 0 else
                           f"Equiv ${c.price:,.2f} > Schein ${r.item.unit_price:,.2f}")
            basis = _basis_text(c)
            if mismatch:
                basis = f"{mismatch}\n{basis}"
            entries.append((rank, order_idx.get(r.item.schein_sku, 999), c.price or 1e9,
                            (r.item, c.scraped_product_name or c.title,
                             c.source_site, c.price, c.url or search_fallback_url(r.item),
                             f"{label} ({match_score(c.criteria, c.confidence)}%)",
                             basis, savings, pct, _source_type_for(c))))
        # catch-all rows: discovered links that were not priced this run
        for c in discovered:
            entries.append((8, order_idx.get(r.item.schein_sku, 999), 1e9,
                            (r.item, c.title or "(discovered listing — not priced)",
                             c.source_site, None, c.url or search_fallback_url(r.item),
                             "NOT PRICED",
                             "DISCOVERED — found during search but not priced this run "
                             "(per-item scrape cap reached). Open the link to check the "
                             "price, or raise SCRAPE_CAP_PER_ITEM to price more per item.",
                             "—", None, _source_type_for(c))))

    entries.sort(key=lambda e: (e[0], e[1], e[2]))
    for band, (_, _, _, args) in enumerate(entries):
        _alt_row(ws, band, *args)
    wb.save(out)
    log.info("Alternate purchases report: %d row(s) → %s", len(entries), out.name)
    return out


# ----------------------------------------------------- evidence report ------

def write_evidence_file(order: ParsedOrder, results: List[ItemResult],
                        findings: List[EquivalencyFinding], out: Path) -> Path:
    """Everything, per PRD 4.8 / 5.6 — every URL, every condition, every
    rejection, every confidence note. Nothing silently discarded."""
    wb = Workbook()

    ws = wb.active
    ws.title = "All Findings"
    title = (f"Background Evidence  ·  Ref: {order.reference or '—'}"
             f"  ·  Generated: {_now()}")
    headers = ["Schein SKU", "Description", "Schein Unit\nPrice", "Candidate Title",
               "Source Site", "Source Type", "Product URL", "Price", "Pack\nQty",
               "Pack/Qty Condition", "Match Type", "Match\nScore", "Confidence",
               "Brand✓", "Name✓", "Size✓", "Pack✓", "Notes / Rejection Reason"]
    widths = [11, 30, 12, 34, 16, 14, 42, 11, 7, 18, 12, 9, 10, 6, 6, 6, 6, 40]
    _title_block(ws, title, "Every price found, every condition, every rejection — nothing discarded",
                 len(headers), headers, widths)
    band = 0
    for r in results:
        all_cands = list(r.candidates) + list(getattr(r, "marketplace_candidates", None) or [])
        if not all_cands:
            ws.append([r.item.schein_sku, r.item.description, r.item.unit_price,
                       "— no public candidates found —", "—", "Other", "—",
                       None, None, "—", "none", 0, 0, "", "", "", "", "—"])
            _style_row(ws, ws.max_row, len(headers), None, band % 2 == 0, {2, 4, 18})
            _money(ws, ws.max_row, [3])
            band += 1
        for c in all_cands:
            ws.append([
                r.item.schein_sku, r.item.description, r.item.unit_price,
                c.title, c.source_site, _source_type_for(c),
                c.url, c.price, c.pack_qty, _dash(_clean(c.pack_condition)),
                c.match_type, match_score(c.criteria, c.confidence), c.confidence,
                *("Y" if c.criteria.get(k) else ("N" if k in c.criteria else "")
                  for k in ("brand_match", "name_match", "size_form_match", "pack_match")),
                c.rejected_reason or _clean(c.notes) or "—",
            ])
            _style_row(ws, ws.max_row, len(headers), None, band % 2 == 0,
                       {2, 4, 10, 18}, link_col=7, url=c.url)
            _money(ws, ws.max_row, [3, 8])
            band += 1

    ws2 = wb.create_sheet("Equivalency Findings")
    h2 = ["Schein SKU", "Original Product", "Equivalent", "Confidence",
          "Basis", "Supplier", "Product URL", "Price", "Routed To"]
    _title_block(ws2, "Equivalency Findings (all levels)", "Table-driven — PRD 5.6",
                 len(h2), h2, [11, 36, 32, 18, 42, 18, 46, 11, 22])
    for band, f in enumerate(findings):
        routed = ("Alternate Purchase List"
                  if f.confidence_level in ("exact_equivalent", "close_equivalent")
                  else "Evidence only — manual review")
        ws2.append([f.item.schein_sku, f.item.description, f.equivalent_name,
                    f.confidence_level.replace("_", " "), f.basis,
                    f.supplier or "—", f.url or "—", f.price, routed])
        _style_row(ws2, ws2.max_row, len(h2), None, band % 2 == 0, {2, 3, 5},
                   link_col=7 if f.url else None, url=f.url or "")
        _money(ws2, ws2.max_row, [8])

    ws3 = wb.create_sheet("Flagged Sites")
    h3 = ["Schein SKU", "Site / URL", "Reason"]
    _title_block(ws3, "Flagged Sites", "Login-required / no public price — skipped, never accessed",
                 len(h3), h3, [11, 50, 38])
    band = 0
    for r in results:
        for s in r.flagged_sites:
            ws3.append([r.item.schein_sku, s, "login required / no public price"])
            _style_row(ws3, ws3.max_row, len(h3), None, band % 2 == 0, {2, 3})
            band += 1
        for c in r.candidates:
            if c.rejected_reason and "login" in c.rejected_reason:
                ws3.append([r.item.schein_sku, c.url, c.rejected_reason])
                _style_row(ws3, ws3.max_row, len(h3), None, band % 2 == 0, {2, 3},
                           link_col=2, url=c.url)
                band += 1

    ws4 = wb.create_sheet("Parsed Order Audit")
    h4 = ["Qty", "Schein SKU", "Description", "UOM", "Unit Price",
          "Extended Price", "Pack Qty", "MPN", "Brand", "Variant"]
    _title_block(ws4, f"Parsed Order Audit  ·  {order.source_file}",
                 "Every extracted line item — totals must reconcile with the PDF",
                 len(h4), h4, [6, 11, 44, 6, 12, 13, 8, 16, 14, 14])
    for band, r in enumerate(results):
        i = r.item
        ws4.append([i.qty, i.schein_sku, i.description, i.uom, i.unit_price,
                    i.extended_price, _dash(i.pack_qty), _dash(i.mpn),
                    _dash(i.brand), _dash(i.variant)])
        _style_row(ws4, ws4.max_row, len(h4), None, band % 2 == 0, {3})
        _money(ws4, ws4.max_row, [5, 6])
    ws4.append([])
    ws4.append(["", "", "Computed total", "", "", order.computed_total])
    ws4.append(["", "", "Printed PDF total", "", "", order.total_price])
    for rr in (ws4.max_row - 1, ws4.max_row):
        ws4.cell(row=rr, column=3).font = BOLD
        ws4.cell(row=rr, column=6).number_format = MONEY
        ws4.cell(row=rr, column=6).font = BOLD

    wb.save(out)
    log.info("Evidence file written → %s", out.name)
    return out


def compute_run_stats(results: List[ItemResult], findings: List[EquivalencyFinding]) -> dict:
    """Summary counters for the results dashboard panel."""
    exact = 0
    near = 0
    no_price = 0
    estimated_savings = 0.0
    for r in results:
        opts = _select_options(r)
        if opts:
            best = opts[0]
            if _is_exact_cand(r.item, best):
                exact += 1
            else:
                near += 1
            if best.price is not None and best.price < r.item.unit_price:
                estimated_savings += (r.item.unit_price - best.price) * r.item.qty
        else:
            no_price += 1
    alt = sum(
        1 for f in findings
        if f.confidence_level in ("exact_equivalent", "close_equivalent")
    )
    return {
        "exact_matches": exact,
        "near_matches": near,
        "alternate_candidates": alt,
        "no_public_price": no_price,
        "estimated_savings": round(estimated_savings, 2),
        "items_processed": len(results),
    }